# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
from datetime import datetime
import hash_data
import crypter
import ast
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

nombre_de_archivo = 'Datos.xlsx'

####    Funciones internas de adicion, supresion y visualizacion           ####
def get_column_letter(column_number):
    dividend = column_number
    column_letter = ''
    modulo = 0

    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_letter = chr(65 + modulo) + column_letter
        dividend = (dividend - modulo) // 26

    return column_letter

def anadir_linea(nombre_cartera, nueva_linea):
    try:
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)
    except FileNotFoundError:
        df = pd.DataFrame()

    #Crea un hash de la linea y lo anade al final de la linea
    hash_value = hash_data.generate_hash(nueva_linea)
    nueva_linea['Hash'] = hash_value

    nueva_linea_df = pd.DataFrame([nueva_linea], columns=df.columns)
    df = pd.concat([df, nueva_linea_df], ignore_index=True)

    with pd.ExcelWriter(nombre_de_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=nombre_cartera, index=False)

def anadir_linea_cliente(nombre_cartera, nueva_linea, key):
    try:
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)
    except FileNotFoundError:
        df = pd.DataFrame()

    hash_value = hash_data.generate_hash(nueva_linea)
    nueva_linea['Hash'] = hash_value
    nueva_linea['Key'] = key

    nueva_linea_df = pd.DataFrame([nueva_linea], columns=df.columns)
    df = pd.concat([df, nueva_linea_df], ignore_index=True)

    with pd.ExcelWriter(nombre_de_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=nombre_cartera, index=False)

def mostrar_lineas(nombre_cartera):
    try:
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)

        print("Contenido de la hoja '{}' del fichero '{}':".format(nombre_cartera, nombre_de_archivo))
        print(df.iloc[:, :-1])
        print('\n')
    except FileNotFoundError:
        print("No se ha encontrado el archivo Excel '{}'.".format(nombre_de_archivo))
    except pd.errors.EmptyDataError:
        print("La hoja '{}' del fichero '{}' est\u00e1 vac\u00eda.".format(nombre_cartera, nombre_de_archivo))

def mostrar_lineas_crypt(nombre_cartera):
    try:
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)

        for index, row in df.iterrows():
            key = ast.literal_eval(row['Key'])
            df.at[index, 'Direccion'] = ast.literal_eval(df.at[index, 'Direccion'])
            df.at[index, 'Telefono'] = ast.literal_eval(df.at[index, 'Telefono'])
            df.at[index, 'Direccion'] = crypter.decrypt_aes(df.at[index, 'Direccion'], key)[2:-1]
            df.at[index, 'Telefono'] = crypter.decrypt_aes(df.at[index, 'Telefono'], key)[2:-1]

        print("Contenido de la hoja '{}' del fichero '{}':".format(nombre_cartera, nombre_de_archivo))
        print(df.iloc[:, :-2])
        print('\n')
    except FileNotFoundError:
        print("No se ha encontrado el archivo Excel '{}'.".format(nombre_de_archivo))
    except pd.errors.EmptyDataError:
        print("La hoja '{}' del fichero '{}' est\u00e1 vac\u00eda.".format(nombre_cartera, nombre_de_archivo))

def borrar_linea(nombre_cartera):
    while True:
        try:
            wb = openpyxl.load_workbook(nombre_de_archivo)
            hoja = wb[nombre_cartera]
            borrar_linea = (input("Introduzca el n\u00famero de la l\u00ednea que desea suprimir (s para salir): "))
            if borrar_linea == 's':
                break
            else:
                borrar_linea = int(borrar_linea)
            hoja.delete_rows(borrar_linea+2)
            wb.save(nombre_de_archivo)
            print("La l\u00ednea {} se ha eliminado correctamente.".format(borrar_linea))
            return (borrar_linea)
        except FileNotFoundError:
            print("No se ha encontrado el archivo Excel '{}.".format(nombre_de_archivo))

def date_heure():
    date_heure = datetime.now()
    date_heure = date_heure.strftime("%Y-%m-%d %H:%M:%S")
    return(date_heure)


####     Funciones de aplicacion        ####
def crear_venta():
    nombre_cartera = 'Ventas'

    #ID Venta
    while True:
        id_venta = int(input("Introduzca el valor de ID Venta (n\u00famero entero): "))
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)
        if id_venta in df['ID Venta'].values:
            print(f"Ya existe una venta con ID Venta {id_venta}. Int\u00e9ntelo de nuevo.")
        else:
            break

    #ID producto
    id_producto = input("Introduzca el valor de ID Producto (n\u00famero): ")
    while not id_producto.isdigit():
        print("ID Producto debe ser un n\u00famero.")
        id_producto = input("Introduzca el valor de ID Producto (n\u00famero): ")

    #Precio
    precio = input("Introduzca el valor de Precio (n\u00famero): ")
    while not precio.isdigit():
        print("Precio debe ser un n\u00famero.")
        precio = input("Introduzca el valor de Precio (n\u00famero): ")

    #Fecha
    while True:
       fecha_str = input("Introduzca el valor de Fecha (DD/MM/YY): ")
       try:
           fecha = datetime.strptime(fecha_str, '%d/%m/%y')
           break
       except ValueError:
           print("Fecha debe estar en el formato DD/MM/YY. Int\u00e9ntelo de nuevo.")

    nueva_linea = {
        'ID Venta': int(id_venta),
        'ID Producto': int(id_producto),
        'Fecha': fecha,
        'Precio': int(precio)
    }
    anadir_linea(nombre_cartera, nueva_linea)
    print("La l\u00ednea se ha a\u00f1adido correctamente a Ventas")
    return(nombre_cartera, nueva_linea)

def modificar_venta():
    nombre_cartera = 'Ventas'
    mostrar_lineas(nombre_cartera)
    wb = openpyxl.load_workbook(nombre_de_archivo)
    hoja = wb[nombre_cartera]
    while True :
        numero_de_linea = input("Introduzca el n\u00famero de la l\u00ednea que desea modificar (s para salir): ")
        if numero_de_linea == "s":
            break
        else:
            numero_de_linea = int(numero_de_linea) + 2

        if numero_de_linea < 1 or numero_de_linea > hoja.max_row:
            print("L\u00ednea no v\u00e1lida.")
            return
        linea_actual = hoja[numero_de_linea]
        print("Contenido actual de la l\u00ednea :")
        for celula in linea_actual[:-1]:
            print(celula.value, end=" | ")
        print("\n")

        try:
            columna_index = input("Introduzca el n\u00famero de la columna que desea modificar : ")
            if columna_index == 's':
                break
            else:
                columna_index = int(columna_index) -1

            if columna_index < 0 or columna_index >= len(hoja[1]):
                print("N\u00famero de columna no v\u00e1lido.")
                return
        except ValueError:
            print("Introduzca un n\u00famero de columna v\u00e1lido.")
            return

        columna = hoja[1][columna_index].value
        letra_columna = get_column_letter(columna_index+1)
        coordonnee = f"{letra_columna}{numero_de_linea}"

        if columna != "Fecha" :
            nouvelle_valeur = input(f"Introduzca el nuevo valor por {columna}: ")
        elif columna == "Fecha" :
            nouvelle_valeur = input(f"Introduzca el nuevo valor por {columna} (DD/MM/YYYY): ")

        #Verifie le format de la nouvelle valeur
        if columna == 'ID Venta' or columna == 'ID Producto':
            try:
                nouvelle_valeur = int(nouvelle_valeur)
            except ValueError:
                print("Ingrese un n\u00famero v\u00e1lido.")
                return
        elif columna == 'Fecha':
            try:
                nouvelle_valeur = datetime.strptime(nouvelle_valeur, '%d/%m/%Y')
            except ValueError:
                print("Ingrese una fecha v\u00e1lida en el formato DD/MM/YYYY.")
                return
        elif columna == 'Precio':
            try:
                nouvelle_valeur = int(nouvelle_valeur)
            except ValueError:
                print("Ingrese un n\u00famero v\u00e1lido para el precio.")
                return

        hoja[coordonnee] = nouvelle_valeur
        linea = {'ID Venta' : hoja[f"A{numero_de_linea}"].value, 'ID Producto': hoja[f"B{numero_de_linea}"].value, 'Fecha': hoja[f"C{numero_de_linea}"].value, 'Precio': hoja[f"D{numero_de_linea}"].value}
    
        #Modifie le hash
        coord_hash = f"E{numero_de_linea}"
        hash_value = hash_data.generate_hash(linea)
        hoja[coord_hash] = hash_value

        #Enregistre
        wb.save(nombre_de_archivo)
        print(f"La l\u00ednea {numero_de_linea-2} en la columna {columna} ha sido modificada con \u00e9xito.")
        return(nombre_cartera, numero_de_linea-2, columna, linea)
        break

def eliminar_venta():
    nombre_cartera = 'Ventas'
    mostrar_lineas(nombre_cartera)
    linea = borrar_linea(nombre_cartera)
    return(nombre_cartera, linea)

def listar_productos():
    nombre_cartera = 'Productos'
    mostrar_lineas(nombre_cartera)

def agregar_productos():
    nombre_cartera = 'Productos'

    #ID Producto
    while True:
        id_prod = int(input("Introduzca el valor de ID Producto (n\u00famero entero): "))
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)
        if id_prod in df['ID Producto'].values:
            print(f"Ya existe una venta con ID Producto {id_prod}. Int\u00e9ntelo de nuevo.")
        else:
            break

    #Nombre
    nombre = input("Introduzca el Nombre : ")

    #Descripcion
    desc = input("Introduzca el Descripcion : ")

    #Precio
    precio = input("Introduzca el valor de Precio (n\u00famero): ")
    while not precio.isdigit():
        print("Precio debe ser un n\u00famero.")
        precio = input("Introduzca el valor de Precio (n\u00famero): ")

    nueva_linea = {
        'ID Producto': int(id_prod),
        'Nombre': nombre,
        'Descripcion': desc,
        'Precio': int(precio)
    }
    anadir_linea(nombre_cartera, nueva_linea)
    print("La l\u00ednea se ha a\u00f1adido correctamente a Productos")
    return(nombre_cartera, nueva_linea)

def desactivar_productos():
    nombre_cartera = 'Productos'
    mostrar_lineas(nombre_cartera)
    linea = borrar_linea(nombre_cartera)
    return(nombre_cartera, linea)

def agregar_usuario():
    nombre_cartera = 'Usuarios'

    #ID Rol
    while True:
        id_rol = int(input("Introduzca el valor de ID Rol (n\u00famero entero): "))
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)
        if id_rol in df['ID Rol'].values:
            print(f"Ya existe una venta con ID Rol {id_rol}. Int\u00e9ntelo de nuevo.")
        else:
            break

    #Nombre
    nombre = input("Introduzca el Nombre : ")

    #Permissions
    perm = input("Introduzca los Permissions : ")

    nueva_linea = {
        'ID Rol': int(id_rol),
        'Nombre': nombre,
        'Permissions': perm,
    }
    anadir_linea(nombre_cartera, nueva_linea)
    print("La l\u00ednea se ha a\u00f1adido correctamente a Usuarios")
    return(nombre_cartera, nueva_linea)

def crear_cliente():
    key = crypter.generate_aes_key()
    nombre_cartera = 'Clientes'

    #Nombre
    nombre = input("Introduzca el valor de Nombre: ")
    
    #Apellidos
    apellidos = input("Introduzca el valor de Apellidos: ")

    #Direccion
    direccion = input("Introduzca el valor de Direccion: ")
    direccion = crypter.encrypt_aes(direccion.encode('utf-8'), key)
    #type : bytes
    
    #Telefono
    telefono = input("Introduzca el valor de Telefono (n\u00famero): ")
    while not telefono.isdigit():
        print("Telefono debe ser un n\u00famero.")
        telefono = input("Introduzca el valor de Telefono (n\u00famero): ")
    telefono = crypter.encrypt_aes(telefono.encode('utf-8'), key)

    nueva_linea = {
        'Nombre': nombre,
        'Apellidos': apellidos,
        'Direccion': direccion,
        'Telefono': telefono
    }
    anadir_linea_cliente(nombre_cartera, nueva_linea, key)
    print("La l\u00ednea se ha a\u00f1adido correctamente a Clientes")
    return(nombre_cartera, nueva_linea, key)

def modificar_cliente():
    while True:
        nombre_cartera = 'Clientes'
        mostrar_lineas_crypt(nombre_cartera)
        df = pd.read_excel(nombre_de_archivo, sheet_name=nombre_cartera)

    #recupere la ligne
        numero_de_linea = (input("Introduzca el n\u00famero de la l\u00ednea que desea modificar (s para salir) : "))
        if numero_de_linea == 's':
            break
        else:
            numero_de_linea = int(numero_de_linea)

        if numero_de_linea < 0 or numero_de_linea > df.shape[0]:
            print("L\u00ednea no v\u00e1lida.")
            return
        print("\nContenido actual de la l\u00ednea :")
        key = ast.literal_eval(df.at[numero_de_linea, 'Key'])
        df.at[numero_de_linea, 'Direccion'] = crypter.decrypt_aes(ast.literal_eval(df.at[numero_de_linea, 'Direccion']), key)[2:-1]
        df.at[numero_de_linea, 'Telefono'] = crypter.decrypt_aes(ast.literal_eval(df.at[numero_de_linea, 'Telefono']), key)[2:-1]
        print(df.iloc[numero_de_linea][:-2])

    #recupere la colonne
        try:
            columna_index = int(input("\nIntroduzca el n\u00famero de la columna que desea modificar : ")) - 1
            if columna_index < 0 or columna_index >= df.shape[1]:
                print("N\u00famero de columna no v\u00e1lido.")
                return
        except ValueError:
            print("Introduzca un n\u00famero de columna v\u00e1lido.")
            return

        columna = df.columns[columna_index]


    #modifie la ligne et recrypte
        nouvelle_valeur = input(f"Introduzca el nuevo valor por {columna}: ")
    
        if columna == 'Telefono':
            while not nouvelle_valeur.isdigit():
                print("Telefono debe ser un n\u00famero.")
                nouvelle_valeur = input("Introduzca el valor de Telefono (n\u00famero): ")
            nouvelle_valeur = crypter.encrypt_aes(nouvelle_valeur.encode('utf-8'), key)
            df.at[numero_de_linea, columna] = nouvelle_valeur
            df.at[numero_de_linea, 'Direccion'] = crypter.encrypt_aes(df.at[numero_de_linea, 'Direccion'].encode('utf-8'), key)
        elif columna == 'Direccion':
            nouvelle_valeur = crypter.encrypt_aes(nouvelle_valeur.encode('utf-8'), key)
            df.at[numero_de_linea, columna] = nouvelle_valeur
            df.at[numero_de_linea, 'Telefono'] = crypter.encrypt_aes(df.at[numero_de_linea, 'Telefono'].encode('utf-8'), key)
        else:
            df.at[numero_de_linea, columna] = nouvelle_valeur
            df.at[numero_de_linea, 'Direccion'] = crypter.encrypt_aes(df.at[numero_de_linea, 'Direccion'].encode('utf-8'), key)
            df.at[numero_de_linea, 'Telefono'] = crypter.encrypt_aes(df.at[numero_de_linea, 'Telefono'].encode('utf-8'), key)

    #change le hash
        hash_enr = df.iloc[numero_de_linea].to_dict()['Hash']
        mon_dict = df.iloc[numero_de_linea].to_dict()
        data = {cle: valeur for index, (cle, valeur) in enumerate(mon_dict.items()) if index < len(mon_dict) - 2}
        nouveau_hash = hash_data.generate_hash(data)
        df.at[numero_de_linea, 'Hash'] = nouveau_hash

    #reenregistre la ligne
        with pd.ExcelWriter(nombre_de_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=nombre_cartera, index=False)

        print(f"La l\u00ednea {numero_de_linea} en la columna {columna} ha sido modificada con \u00e9xito.")
        return(nombre_cartera, numero_de_linea, columna, df)
        break

def eliminar_cliente():
    nombre_cartera = 'Clientes'
    mostrar_lineas_crypt(nombre_cartera)
    linea = borrar_linea(nombre_cartera)
    return(nombre_cartera, linea)


####    Fonction d'execution    ####
def exec(nom_action, funcione_usuario):
    date = date_heure()
    act = funcione_usuario

    if nom_action == 'Crear venta':
        ligne = crear_venta()
        obj = "Hoja : " + ligne[0]
        nueva_linea = ligne[1]
        nueva_linea['Fecha'] = nueva_linea['Fecha'].strftime("%Y-%m-%d %H:%M:%S")
        nueva_linea = dict(list(nueva_linea.items())[:4])
        nueva_linea = json.dumps(nueva_linea)
        obj =  obj + ", Nueva linea : " + nueva_linea
    elif nom_action == 'Modificar venta' :
        ligne = modificar_venta()
        obj = "hoja : " + ligne[0] + ", l\u00ednea :" + str(ligne[1]) + ", columna :" + str(ligne[2])
    elif nom_action == 'Eliminar venta' :
        ligne = eliminar_venta()
        obj = "hoja : " + ligne[0] + ", l\u00ednea :" + str(ligne[1])
    elif nom_action == 'Listar productos' :
        listar_productos()
        obj = "hoja : Productos"
    elif nom_action == 'Agregar productos' :
        ligne = agregar_productos()
        obj = "hoja : " + ligne[0]
        nueva_linea = ligne[1]
        nueva_linea = dict(list(nueva_linea.items())[:4])
        nueva_linea = json.dumps(nueva_linea)
        obj =  obj + ", Nueva linea : " + nueva_linea
    elif nom_action == 'Desactivar productos' :
        ligne = desactivar_productos()
        obj = "hoja : " + ligne[0] + ", l\u00ednea :" + str(ligne[1])
    elif nom_action == 'Agregar usuario':
        ligne = agregar_usuario()
        obj = "Hoja : " + ligne[0]
        nueva_linea = ligne[1]
        nueva_linea = dict(list(nueva_linea.items())[:3])
        nueva_linea = json.dumps(nueva_linea)
        obj =  obj + ", Nueva linea : " + nueva_linea
    elif nom_action == 'Crear cliente':
        ligne = crear_cliente()
        obj = "Hoja : " + ligne[0]
        nueva_linea = ligne[1]
        key = nueva_linea['Key']
        nueva_linea['Direccion'] = crypter.decrypt_aes(nueva_linea['Direccion'], key)[2:-1]
        nueva_linea['Telefono'] = crypter.decrypt_aes(nueva_linea['Telefono'], key)[2:-1]
        nueva_linea = dict(list(nueva_linea.items())[:4])
        nueva_linea = json.dumps(nueva_linea)
        obj =  obj + ", Nueva linea : " + nueva_linea
    elif nom_action == 'Modificar cliente':
        ligne = modificar_cliente()
        obj = "hoja : " + ligne[0] + ", l\u00ednea :" + str(ligne[1]) + ", columna :" + str(ligne[2])
    elif nom_action == 'Eliminar cliente':
        ligne = eliminar_cliente()
        obj = "hoja : " + ligne[0] + ", l\u00ednea :" + str(ligne[1])

    #Ajoute la ligne d action a  audit.txt
    nvl_ligne = date + " Accion : " + nom_action + ", " + obj + " por " + act + "\n"
    with open('Audit.txt', 'a') as fichier:
        fichier.write(nvl_ligne)