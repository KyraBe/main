import hashlib
import json
from datetime import date
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import shutil
import base64
import ast
import modifier_donnees

nom_fichier='Datos.xlsx'

###         Fonctions internes      ###
def recuperer_ligne(nom_fichier, nom_feuille, numero_ligne):
    try:
        wb = openpyxl.load_workbook(nom_fichier)
        feuille = wb[nom_feuille]
        
        # Accedez a la ligne specifiee
        ligne = feuille.iter_rows(min_row=numero_ligne+2, max_row=numero_ligne+2, values_only=True)
        ligne_data = next(ligne, None)
        
        if ligne_data:
            return ligne_data
        else:
            print(f"Aucune donnee trouvee pour la ligne {numero_ligne} dans la feuille {nom_feuille}")

    except FileNotFoundError:
        print(f"Le fichier Excel '{nom_fichier}' n a pas ete trouve.")
    except Exception as e:
        print(f"Une erreur s est produite : {e}")

def custom_encoder(obj):
    if isinstance(obj, date):
        return obj.isoformat()
    if isinstance(obj, bytes):
        return base64.b64encode(obj).decode('utf-8')
    raise TypeError("Type {type(obj)} not serializable")


###         Fonction pour generer le hash       ###
def generate_hash(data):
    try:
        data_str = json.dumps(data, sort_keys=True, default=custom_encoder)
    except Exception as e:
        print(f"Error while serializing data: {e}")
        print(f"Offending data: {data}")
        raise
    hash_object = hashlib.sha256(data_str.encode())
    return hash_object.hexdigest()

def supprimer_ligne_bis(nom_feuille, numero_ligne):
    try:
        wb = openpyxl.load_workbook(nom_fichier)
        feuille = wb[nom_feuille]
        feuille.delete_rows(numero_ligne+2)
        wb.save(nom_fichier)
    except FileNotFoundError:
        print("No se ha encontrado el archivo Excel '{}.".format(nom_fichier))


####        Fonctions de choix si problemes d'integrite pour l'utilisateur       ####
def eliminer_donnee(df, numero_ligne, nom_feuille):
    excel_file = pd.ExcelFile(nom_fichier)
    df = excel_file.parse(nom_feuille)
    df.drop(numero_ligne, inplace=True)
    with pd.ExcelWriter('Datos.xlsx') as writer:
        df.to_excel(writer, sheet_name=nom_feuille, index=False)
        for sheet_name in excel_file.sheet_names:
            if sheet_name != nom_feuille:
                original_df = excel_file.parse(sheet_name)
                original_df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"\nSe ha eliminado la l\u00ednea {numero_ligne} de la hoja {nom_feuille}")

def cloner_et_marquer():
    nouveau_nom = f"{nom_fichier}_Perdida_Integridad.xlsx"
    shutil.copy2(nom_fichier, nouveau_nom)
    print(f"El documento {nom_fichier} ha sido clonado y marcado como 'P\u00e9rdida de integridad'.")

def modifier_et_regenerer(nom_feuille, numero_ligne):
    ligne_modifiee = recuperer_ligne(nom_fichier, nom_feuille, numero_ligne)
    supprimer_ligne_bis(nom_feuille, numero_ligne)
    excel_file = pd.ExcelFile(nom_fichier)
    df = excel_file.parse(nom_feuille)

    wb = openpyxl.load_workbook(nom_fichier)
    feuille = wb[nom_feuille]

    noms_colonnes = df.columns.tolist()
    print(f"\nLigne actuelle")
    for colonne, valeur in zip(noms_colonnes, ligne_modifiee):
        print(f"{colonne}: {valeur}", end=" | ")

    try:
        colonne_index = int(input("\nIntroduzca el n\u00famero de la columna que desea modificar : ")) - 1
        if colonne_index < 0 or colonne_index >= len(feuille[1]):
            print("N\u00famero de columna no v\u00e1lido.")
            return
    except ValueError:
        print("Introduzca un n\u00famero de columna v\u00e1lido.")
        return

    colonne = feuille[1][colonne_index].value
    colonne_lettre = get_column_letter(colonne_index+1)
    coordonnee = f"{colonne_lettre}{numero_ligne+2}"

    nouvelle_valeur = input("Introduzca el nuevo valor : ")

    #verifie le format de la nouvelle valeur
    if colonne == 'ID Venta' or colonne == 'ID Producto' or colonne == 'ID Rol':
        try:
            nouvelle_valeur = int(nouvelle_valeur)
        except ValueError:
            print("Ingrese un n\u00famero v\u00e1lido.")
            return
    elif colonne == 'Fecha':
        try:
            nouvelle_valeur = datetime.strptime(nouvelle_valeur, '%d/%m/%Y').date()
        except ValueError:
            print("Ingrese una fecha v\u00e1lida en el formato DD/MM/YYYY.")
            return
    elif colonne == 'Precio':
        try:
            nouvelle_valeur = float(nouvelle_valeur)
        except ValueError:
            print("Ingrese un n\u00famero v\u00e1lido para el precio.")
            return

    #modifie la ligne et en ajoute une nouvelle
    feuille[coordonnee] = nouvelle_valeur
    print(f"La l\u00ednea {numero_ligne} en la columna {colonne} ha sido modificada con \u00e9xito.")
    ligne_modifiee = list(ligne_modifiee)
    ligne_modifiee[colonne_index] = nouvelle_valeur
    hash_column = ligne_modifiee.pop()

    #Genere le hash et l ajoute a la ligne
    hash_modifie = generate_hash(ligne_modifiee)
    ligne_modifiee.append(hash_column)
    ligne_modifiee[-1] = hash_modifie
    ligne_modifiee_df = pd.DataFrame([ligne_modifiee], columns=df.columns)
    df = pd.concat([df, ligne_modifiee_df], ignore_index=True)

    #sauvegarde
    wb.save(nom_fichier)
    with pd.ExcelWriter(nom_fichier, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=nom_feuille, index=False)


####        Fonctions pour check l'intergrite et faire un choix     ####
def check_integrity(funcione_usuario):
    nom_fichier='Datos.xlsx'
    wb = openpyxl.load_workbook(nom_fichier)
    problemes_integrite = []

    #cherche tous les problemes d'integrite
    for nom_feuille in wb.sheetnames:
        feuille = wb[nom_feuille]

        for numero_ligne in range(2, feuille.max_row + 1):
            ligne_actuelle = feuille[numero_ligne]

            if nom_feuille != "Clientes" :
                hash_enregistre = ligne_actuelle[feuille.max_column-1].value
                donnees_ligne = {feuille[1][colonne].value: ligne_actuelle[colonne].value for colonne in range(0, feuille.max_column-1)}
            elif nom_feuille == "Clientes" :
                #exception si c'est la feuille Clientes car des donnees sont en bytes
                hash_enregistre = ligne_actuelle[feuille.max_column-2].value
                donnees_ligne = {feuille[1][colonne].value: ligne_actuelle[colonne].value for colonne in range(0, feuille.max_column-2)}
                donnees_ligne['Direccion'] = ast.literal_eval(donnees_ligne['Direccion'])
                donnees_ligne['Telefono'] = ast.literal_eval(donnees_ligne['Telefono'])

            nouveau_hash = generate_hash(donnees_ligne)

            #verifie si le hash est bien le meme
            if hash_enregistre != nouveau_hash:
                problemes_integrite.append((nom_feuille, numero_ligne-2))

    if problemes_integrite:
        #si il y a des problemes d'integrite, l'utilisateur choisit quoi faire
        date_heure = modifier_donnees.date_heure()
        cat = " Problemas de integridad "
        obj = []
        for nom_feuille, numero_ligne in problemes_integrite:
            obj.append((nom_feuille, numero_ligne))
        
        nvl_ligne = date_heure + cat + " ".join(map(str, obj)) + "\n"
        with open('Audit.txt', 'a') as fichier:
            fichier.write(nvl_ligne)

        choix_integrite(problemes_integrite, funcione_usuario)
    else:
        print("No se detectan problemas de integridad.")

def choix_integrite(problemes_integrite, funcione_usuario):
    choix = []
    for (nom_feuille, numero_ligne) in problemes_integrite:
        print(f"\nProblema de integridad en la p\u00e1gina {nom_feuille} en la l\u00ednea {numero_ligne}:")
        df = pd.read_excel('Datos.xlsx', sheet_name=nom_feuille)
        print(df.iloc[numero_ligne]) 

        print("\nOpciones:")
        print("1. Borrar dato")
        print("2. Clonar y marcar como 'P\u00e9rdida de integridad'")
        print("3. Modificar los datos y restaurar la clave de integridad")

        choix_utilisateur = input("\nElija una opci\u00f3n (1, 2, 3) u otra para no hacer nada : ")
        choix.append((nom_feuille, numero_ligne, choix_utilisateur))

    for (nom_feuille, numero_ligne, choix_utilisateur) in reversed(choix):
        df = pd.read_excel('Datos.xlsx', sheet_name=nom_feuille)
        date_heure = modifier_donnees.date_heure()
        cat = " Integridad corrupta "
        act = funcione_usuario

        if choix_utilisateur == "1":
            eliminer_donnee(df, numero_ligne, nom_feuille)  
            accion = "Borrar l\u00ednea " + str(numero_ligne) + " p\u00e1gina " + nom_feuille
        elif choix_utilisateur == "2":
            cloner_et_marquer()
            accion = "Clonaci\u00f3n y etiquetado de archivo"
        elif choix_utilisateur == "3":
            modifier_et_regenerer(nom_feuille, numero_ligne)
            accion = "Modificac\u00f3n la hoja " + nom_feuille + " l\u00ednea " + str(numero_ligne )+ " y regeneraci\u00f3n de hash"
        else:
            print("Opci\u00f3n no v\u00e1lida. Pase al siguiente problema.")

        #ajoute une ligne a Audit.txt
        nvl_ligne = date_heure + cat + accion + " por " + act + "\n"
        with open('Audit.txt', 'a') as fichier:
            fichier.write(nvl_ligne)
